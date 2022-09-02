import shutil
import sys
from datetime import datetime
from time import sleep
import PyPDF2
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from easygui import *


caminho_navarro = r'C:\Users\Administrador\Desktop\NAVARRO'
ponto_inicial = Path.cwd()


def consulta_procuracoes():
    print('Consultado procurações pendentes de ajuizamento...')
    pasta_pendentes = Path(caminho_navarro + r'\DOCUMENTACAO PENDENTE')
    procuracoes = []
    for procuracao in pasta_pendentes.iterdir():
        procuracao = procuracao.stem
        procuracoes.append(procuracao.split('-')[-1])
    print('Procurações verificadas.')
    return procuracoes


def consulta_credenciais(pessoa):
    print('Consultando credenciais do portal da transparência...')
    credenciais = pd.read_excel(rf'{caminho_navarro}\PARAMETROS' + r'\relatorio_vinculos.xlsx')
    matricula = credenciais[credenciais.Cliente == pessoa].Matricula.values[0]
    login = credenciais[credenciais.Cliente == pessoa].Login.values[0]
    senha = credenciais[credenciais.Cliente == pessoa].Senha.values[0]

    while 1:
        if matricula != '' and login != '' and senha != '':
            break
        else:
            msg = "Informe as credenciais do/a cliente"
            title = "Credenciais"
            nome_campos = ["Matrícula: ", "Login: ", "Senha: "]
            valor_campos = []  # CRIA VETOR PARA RECEBER RESPOSTAS
            valor_campos = multenterbox(msg, title, nome_campos)

            # GARANTE QUE NENHUM CAMPO FIQUE VAZIO
            while 1:
                if valor_campos is None:
                    break
                errmsg = ""
                for i in range(len(nome_campos)):
                    if valor_campos[i].strip() == "":
                        errmsg = errmsg + ('"%s" é um campo obrigatório.\n\n' % nome_campos[i])
                if errmsg == "":
                    break  # no problems found
                valor_campos = multenterbox(errmsg, title, nome_campos, valor_campos)
    print('Credenciais obtidas.')

    return matricula, login, senha


def prepara_modelo_calculo():
    print('Preparanto planilha para cálculo automatizado...')
    # PREPARAÇÃO PLANILHA MODELO DE CÁLCULO
    modelo = load_workbook(filename=rf'{pasta_parametros}\Calculo Hora Atividade - MODELO.xlsm',
                           read_only=False, keep_vba=True)
    ws = modelo['Capa']
    ws['D3'] = str(pasta_fichas)
    ws['D4'] = f'fichas_financeiras {matricula}'
    ws['D7'] = ano_inicial_estatico
    ws['D8'] = ano_final
    ws['D10'] = rf'{caminho_navarro}\CLIENTES\{cliente}\HORA ATIVIDADE\Calculo {matricula}'
    ws['D10'] = cliente
    ws['D41'] = fr'{caminho_navarro}\PARAMETROS'

    modelo.save(str(pasta_cumprimento) + r'\Memoria de Calculo.xlsm')
    print('Modelo de cálculo preparado.')


def acessa_pagina(login, senha):
    # ACESSO PÁGINA
    link = 'https://senior.blumenau.sc.gov.br/restrito/login.htm'
    navegador.get(url=link)

    # ACESSO LOGIN
    navegador.find_element(by=By.XPATH, value='//*[@id="txtNomUsu"]').send_keys(login)
    navegador.find_element(by=By.XPATH, value='//*[@id="txtSenUsu"]').send_keys(senha)
    navegador.find_element(by=By.XPATH, value='//*[@id="submit"]').click()
    navegador.switch_to.window(navegador.window_handles[1])

    consulta_janelas = len(navegador.window_handles)
    while consulta_janelas < 3:
        try:
            print('Testando credenciais...')
            WebDriverWait(navegador, 20).until(EC.alert_is_present())
            alert = navegador.switch_to.alert
            alert_text = alert.text
            print(alert_text)
            alert.accept()
            print("Falha no login")
            falha = True
            break
        except TimeoutException:
            print("Possível login...")
            consulta_janelas = len(navegador.window_handles)
            falha = False

        if falha:
            continue
    print('Login efetuado!')
    print('Organizando navegadores...')
    sleep(5)
    navegador.window_handles[2]

    # FECHA O NOVO NAVEGADOR
    navegador.close()

    # TROCA GUIA PARA PDF
    navegador.switch_to.window(navegador.window_handles[0])


def consulta_ficha_funcional(matricula, login):
    link_funcional = f'https://senior.blumenau.sc.gov.br/restrito/conector?' \
                     f'ACAO=EXEREL&SIS=FP&NOME=FPDO501.COL&order_Detalhe_1=' \
                     f'0;&dado_EABREMP={login[0]}&dado_EABRCAD={matricula}'

    navegador.switch_to.window(navegador.window_handles[0])
    navegador.get(link_funcional)
    sleep(5)


def consulta_demonstrativo():
    link_demonstrativo = 'https://senior.blumenau.sc.gov.br/restrito/conector?ACAO=EXECUTAREGRA&SIS=FP&REGRA=306'
    navegador.switch_to.window(navegador.window_handles[0])
    navegador.get(link_demonstrativo)
    admissao = navegador.find_element(by=By.XPATH,
                                      value='/html/body/table/tbody/tr/td/table/'
                                            'tbody/tr[3]/td/table/tbody/tr[6]/td[2]').text[-4:]
    desligamento = navegador.find_element(by=By.XPATH,
                                          value='/html/body/table/tbody/tr/td/table/'
                                                'tbody/tr[3]/td/table/tbody/tr[6]/td[4]').text[6:10]
    return int(admissao), int(desligamento)


# def le_ficha_funcional():
#     # Abre o arquivo pdf
#     pdf_file = open(caminho_navarro + r'\DOWNLOADS\relatorio.pdf', 'rb')
#
#     # Faz a leitura usando a biblioteca
#     read_pdf = PyPDF2.PdfFileReader(pdf_file, strict=False)
#
#     # lê a primeira página completa
#     page = read_pdf.getPage(0)
#
#     # extrai apenas o texto
#     page_content = page.extractText()
#
#     len(page_content)
#     # faz a junção das linhas
#     parsed = ''.join(page_content)
#
#     # remove as quebras de linha
#     # parsed = re.sub('\n', '', parsed)
#     with open('texto_conteudo.txt', 'w') as w:
#         w.write(parsed)
#
#     # TRATAMENTO DA NOMEAÇÃO
#     posterior_ato = parsed.split('NOMEAÇÃO')[1]
#     posse = posterior_ato.split('POSSE')[0]
#     data_nomeacao = posse.split('DATA')[1]
#     resultado_data_nomeacao = data_nomeacao[1:]
#     print('NOMEACAO: ', resultado_data_nomeacao)
#
#     # TRATAMENTO DO DESLIGAMENTO
#     posterior_desligamento = parsed.split('DATA DO DESLIGAMENTO')[1][2:11]
#     print('DESLIGAMENTO: ', posterior_desligamento)
#
#     # TRATAMENTO DO VÍNCULO
#     posterior_vinculo = parsed.split('VÍNCULO')[1]
#     anterior_situacao = posterior_vinculo.split('SITUAÇÃO')[0]
#     resultado_vinculo = anterior_situacao[2:]
#     print('VINCULO: ', resultado_vinculo)
#
#     return resultado_data_nomeacao, posterior_desligamento, resultado_vinculo


def organiza_pasta_cumprimento():
    print('Adicionando documentos comuns na pasta de cumprimento de sentença...')
    shutil.copyfile(str(pasta_fichas) + fr"\fichas_financeiras {matricula}.pdf",
                    str(pasta_cumprimento) + rf'\5 Fichas financeiras.pdf')

    shutil.copyfile(str(pasta_parametros) + fr"\4 Titulo executivo judicial - Coletiva.pdf",
                    str(pasta_cumprimento) + rf'\4 Titulo executivo judicial - Coletiva.pdf')

    shutil.copyfile(str(pasta_parametros) + fr"\7 Substabelecimento.pdf",
                    str(pasta_cumprimento) + rf'\7 Substabelecimento.pdf')
    print('Documentos inseridos!')


# ITERAÇÃO POR MATRÍCULA
for cliente in consulta_procuracoes():
    chromeOptions = webdriver.ChromeOptions()
    prefs = {"download.default_directory": rf"{caminho_navarro}\DOWNLOADS",
             "download.prompt_for_download": False,  # To auto download the file
             "download.directory_upgrade": True,
             "plugins.always_open_pdf_externally": True,  # It will not show PDF directly in chrome
             "plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}],
             "download.extensions_to_open": "applications/pdf"
             }
    chromeOptions.add_experimental_option("prefs", prefs)
    navegador = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chromeOptions)
    navegador.implicitly_wait(30)
    actions = ActionChains(navegador)

    print('*' * 20)
    print(cliente)

    # ATRIBUTOS
    matricula, login, senha = consulta_credenciais(cliente)
    acessa_pagina(login, senha)
    # consulta_ficha_funcional(matricula, login)
    # nomeacao, desligamento, situacao = le_ficha_funcional()
    consulta_demonstrativo()
    ano_inicial, ano_final = consulta_demonstrativo()

    if ano_inicial < 2013:
        ano_inicial = 2013
        ano_inicial_estatico = 2013
    else:
        ano_inicial_estatico = ano_inicial

    if ano_final == '':
        ano_final = 2022

    periodo = ano_final - ano_inicial + 1
    print('Ano inicial: ', ano_inicial, 'ANo final: ', ano_final, 'Período: ', periodo)

    # CRIA PASTA DE FICHAS
    pasta_fichas = rf'{caminho_navarro}\CLIENTES\{cliente}\HORA ATIVIDADE\Fichas {matricula}'
    Path.mkdir(Path(pasta_fichas), parents=True)
    pasta_cumprimento = rf'{caminho_navarro}\CLIENTES\{cliente}\HORA ATIVIDADE\Cumprimento de sentença'
    Path.mkdir(Path(pasta_cumprimento))
    pasta_parametros = rf'{caminho_navarro}\PARAMETROS'

    # ACESSO E DOWNLOAD FICHA FINANCEIRA
    mes_atual = '12'
    for ano in range(0, periodo):
        if ano_inicial == '2022':
            data = datetime.now()
            mes_atual = int(data.strftime("%m"))
            mes_atual -= 1
            if mes_atual < 10:
                mes_atual = '0' + str(mes_atual)
                str(mes_atual)
            else:
                str(mes_atual)

        print(f'Mês atual: {mes_atual} | Ano inicial {ano_inicial}')
        ano_inicial = str(ano_inicial)

        link_financeiro = 'https://senior.blumenau.sc.gov.br/restrito/conector?ACAO=EXEREL&SIS=FP&NOME=FPFF512' \
                          '.OPE&dado_EMosUsu=S&dado_ENTipCal=3&dado_ECalcMed=S&dado_EspNivTot=0&dado_EspNivQue=0&' \
                          'dado_ENSomTot=N&dado_EAbrEmp=' + login[0] + '&dado_EAbrTcl=' + login[2] + '&dado_EAb' \
                                                                                                     'rCad=' + \
                          str(matricula) + '&dado_EAbrTco=1-99&dado_EAbrTsa=1-99&dado_EAbrSit=1-999&dado_EAbr' \
                                           'Eve=1-9999&dado_EAbrNEv=0&LINWEB=&dado_EL' \
                                           'istarRef=S&dado_ENPerIni=01/' + str(ano_inicial) + \
                          '&dado_EDatRef=' + mes_atual + '/' + str(ano_inicial)
        navegador.get(url=link_financeiro)

        try:
            print('Confere período...')
            WebDriverWait(navegador, 20).until(EC.alert_is_present())
            alert = navegador.switch_to.alert
            alert_text = alert.text
            print(alert_text)
            alert.accept()
            print("Contracheque indisponível! Provável aposentadoria ou exoneração.")
            falha_periodo = True
        except TimeoutException:
            print("Contracheque disponível!")
            falha_periodo = False

        if falha_periodo:
            continue

        sleep(3)

        pasta_downloads = caminho_navarro + r"\Downloads"
        shutil.move(str(pasta_downloads) + fr"\relatorio.pdf",
                    pasta_fichas + rf'\ficha financeira {matricula} {ano_inicial}.pdf')

        ano_inicial = int(ano_inicial)
        ano_inicial += 1
        ano_inicial = str(ano_inicial)

    # UNIFICA OS PDFS
    from junta_pdf import juntaPDF

    juntaPDF(matricula, ponto_inicial, pasta_fichas)

    # COPIA FICHAS CONSOLIDADAS PARA PASTA DE CUMPRIMENTO
    shutil.copyfile(str(pasta_fichas) + fr"\fichas_financeiras {matricula}.pdf",
                    str(pasta_cumprimento) + rf'\5 Fichas financeiras.pdf')

    # CONVERTE O PDF CONSOLIDADO EM EXCEL
    from adobeSimples import conversao_excel

    conversao_excel(navegador, matricula, pasta_downloads, pasta_fichas)

    organiza_pasta_cumprimento()
    prepara_modelo_calculo()

    p = Path(caminho_navarro + r'\DOCUMENTACAO PENDENTE')
    shutil.move([x for x in p.iterdir() if x.is_file() if x.stem.split('-')[-1] == cliente][0],
                pasta_cumprimento + r'\6 Procuração.pdf')


# FECHA O NAVEGADOR
navegador.quit()
msgbox('Pastas e contracheques preparados')
